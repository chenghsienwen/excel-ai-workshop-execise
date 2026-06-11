"""Unit tests for src/loader.py."""
import pathlib

import pytest

from src import config
from src.loader import (
    _col_b,
    _expand_data_row,
    _make_record,
    _parse_filename,
    _parse_rows,
    _scan_step,
    load_excel,
)


# ── helpers ──────────────────────────────────────────────────────────────────

def _header_row(label="Gross Sales ($K)"):
    """Minimal worksheet row whose col B is a rev_op_type header."""
    return (None, label) + (None,) * 42


def _data_row(region="APAC", amounts=None):
    """Worksheet row with a region in col B and 12 Actual-section amounts."""
    row = [None] * 44
    row[1] = region
    for i, v in enumerate(amounts or [float(i + 1) for i in range(12)]):
        row[2 + i] = v
    return tuple(row)


# ── _parse_filename ───────────────────────────────────────────────────────────

class TestParseFilename:
    def test_cde_product_and_year(self):
        assert _parse_filename(pathlib.Path("CDE biz status_2024_v1.xlsx")) == ("CDE", 2024)

    def test_pj_product_and_year(self):
        assert _parse_filename(pathlib.Path("PJ biz status_2025_v2.xlsx")) == ("PJ", 2025)

    def test_product_always_uppercased(self):
        product, _ = _parse_filename(pathlib.Path("cde biz status_2024_v1.xlsx"))
        assert product == "CDE"

    def test_year_extracted_correctly(self):
        _, year = _parse_filename(pathlib.Path("CDE biz status_2030_v1.xlsx"))
        assert year == 2030

    def test_invalid_filename_raises_value_error(self):
        with pytest.raises(ValueError, match="Cannot infer"):
            _parse_filename(pathlib.Path("unknown_file.xlsx"))

    def test_non_numeric_year_raises(self):
        with pytest.raises(ValueError):
            _parse_filename(pathlib.Path("CDE biz status_YEAR_v1.xlsx"))

    def test_error_message_contains_filename(self):
        name = "bad_report.xlsx"
        with pytest.raises(ValueError, match=name):
            _parse_filename(pathlib.Path(name))


# ── _col_b ────────────────────────────────────────────────────────────────────

class TestColB:
    def test_normal_string_stripped(self):
        assert _col_b((None, "  APAC  ")) == "APAC"

    def test_float_nan_returns_none(self):
        assert _col_b((None, float("nan"))) is None

    def test_none_value_returns_none(self):
        assert _col_b((None, None)) is None

    def test_integer_converted_to_string(self):
        assert _col_b((None, 42)) == "42"

    def test_float_value_converted_to_string(self):
        assert _col_b((None, 1.5)) == "1.5"

    def test_row_too_short_returns_none(self):
        assert _col_b(("only_col_a",)) is None

    def test_empty_row_returns_none(self):
        assert _col_b(()) is None

    def test_whitespace_only_stripped_to_empty_string(self):
        assert _col_b((None, "   ")) == ""

    def test_tab_and_newline_stripped(self):
        assert _col_b((None, "\tNet\n")) == "Net"


# ── _scan_step ────────────────────────────────────────────────────────────────

class TestScanStep:
    def test_none_col_b_keeps_rev_op_and_skips_row(self):
        assert _scan_step(("Gross", None), (None, None)) == ("Gross", None)

    def test_gross_header_sets_rev_op(self):
        rev_op, row = _scan_step((None, None), (None, "Gross Sales ($K)"))
        assert rev_op == "Gross"
        assert row is None

    def test_net_header_updates_rev_op(self):
        rev_op, _ = _scan_step(("Gross", None), (None, "Net Sales ($K)"))
        assert rev_op == "Net"

    def test_op_header_updates_rev_op(self):
        rev_op, _ = _scan_step(("Gross", None), (None, "OP Sales ($K)"))
        assert rev_op == "Op"

    @pytest.mark.parametrize("label", ["Actual", "Forecast", "Budget"])
    def test_section_type_label_skips_row(self, label):
        result = _scan_step(("Gross", None), (None, label))
        assert result == ("Gross", None)

    @pytest.mark.parametrize("label", ["ACTUAL", "FORECAST", "BUDGET"])
    def test_section_type_label_case_insensitive(self, label):
        result = _scan_step(("Gross", None), (None, label))
        assert result == ("Gross", None)

    def test_data_row_returned_with_current_rev_op(self):
        row = (None, "APAC", 100.0)
        assert _scan_step(("Gross", None), row) == ("Gross", row)

    def test_data_row_skipped_when_no_rev_op_seen_yet(self):
        row = (None, "APAC", 100.0)
        assert _scan_step((None, None), row) == (None, None)

    def test_rev_op_carries_forward_over_empty_rows(self):
        state = ("Net", None)
        state = _scan_step(state, (None, None))
        state = _scan_step(state, (None, None))
        assert state[0] == "Net"

    def test_new_header_overrides_previous_rev_op(self):
        state = _scan_step(("Gross", None), (None, "Net Sales ($K)"))
        assert state[0] == "Net"


# ── _make_record ──────────────────────────────────────────────────────────────

_ACTUAL_SECTION = ("Actual", 1, 2)


class TestMakeRecord:
    def test_basic_record_fields(self):
        record = _make_record("CDE", 2024, "Gross", _data_row(), _ACTUAL_SECTION, (0, "Jan"))
        assert record == {
            "product": "CDE",
            "year": 2024,
            "region": "APAC",
            "rev_op_type": "Gross",
            "sales_budget_type": "Actual",
            "month": "Jan",
            "amount": 1.0,
        }

    def test_none_region_returns_none(self):
        row = list(_data_row())
        row[1] = None
        assert _make_record("CDE", 2024, "Gross", tuple(row), _ACTUAL_SECTION, (0, "Jan")) is None

    def test_region_col_out_of_bounds_returns_none(self):
        assert _make_record("CDE", 2024, "Gross", (None,), _ACTUAL_SECTION, (0, "Jan")) is None

    def test_amount_none_becomes_zero(self):
        row = list(_data_row())
        row[2] = None  # Jan amount (month_start=2, m_idx=0)
        record = _make_record("CDE", 2024, "Gross", tuple(row), _ACTUAL_SECTION, (0, "Jan"))
        assert record["amount"] == 0.0

    def test_amount_col_out_of_bounds_becomes_zero(self):
        # Row only has region; month column is out of bounds
        short_row = (None, "APAC")
        record = _make_record("CDE", 2024, "Gross", short_row, _ACTUAL_SECTION, (0, "Jan"))
        assert record["amount"] == 0.0

    def test_region_stripped(self):
        row = list(_data_row(region="  APAC  "))
        record = _make_record("CDE", 2024, "Gross", tuple(row), _ACTUAL_SECTION, (0, "Jan"))
        assert record["region"] == "APAC"

    def test_month_index_selects_correct_amount(self):
        amounts = [float(m) for m in range(1, 13)]
        row = _data_row(amounts=amounts)
        for i, month_name in enumerate(config.MONTH_COLS):
            record = _make_record("CDE", 2024, "Gross", row, _ACTUAL_SECTION, (i, month_name))
            assert record["amount"] == float(i + 1)
            assert record["month"] == month_name

    @pytest.mark.parametrize("section", config.SECTION_CONFIG)
    def test_all_sections_produce_valid_records(self, section, full_data_row):
        record = _make_record("CDE", 2024, "Gross", full_data_row, section, (0, "Jan"))
        assert record is not None
        assert record["sales_budget_type"] == section[0]


# ── _expand_data_row ──────────────────────────────────────────────────────────

class TestExpandDataRow:
    def test_full_row_produces_all_section_month_combos(self, full_data_row):
        records = _expand_data_row("CDE", 2024, "Gross", full_data_row)
        assert len(records) == len(config.SECTION_CONFIG) * len(config.MONTH_COLS)

    def test_partial_row_filters_missing_regions(self):
        # Only the Actual region (col 1) is set; Forecast and Budget are None
        row = [None] * 44
        row[1] = "APAC"
        for i in range(12):
            row[2 + i] = float(i + 1)
        records = _expand_data_row("CDE", 2024, "Gross", tuple(row))
        assert len(records) == 12
        assert all(r["sales_budget_type"] == "Actual" for r in records)

    def test_product_year_rev_op_embedded(self, full_data_row):
        records = _expand_data_row("PJ", 2025, "Net", full_data_row)
        assert all(r["product"] == "PJ" for r in records)
        assert all(r["year"] == 2025 for r in records)
        assert all(r["rev_op_type"] == "Net" for r in records)

    def test_all_months_covered(self, full_data_row):
        records = _expand_data_row("CDE", 2024, "Gross", full_data_row)
        months = {r["month"] for r in records}
        assert months == set(config.MONTH_COLS)

    def test_empty_row_returns_no_records(self):
        records = _expand_data_row("CDE", 2024, "Gross", ())
        assert records == []


# ── _parse_rows ───────────────────────────────────────────────────────────────

class TestParseRows:
    def test_empty_input_returns_empty_list(self):
        assert _parse_rows([], "CDE", 2024) == []

    def test_data_row_before_header_is_skipped(self):
        assert _parse_rows([_data_row()], "CDE", 2024) == []

    def test_header_then_data_produces_records(self):
        records = _parse_rows([_header_row(), _data_row()], "CDE", 2024)
        assert len(records) > 0

    def test_header_sets_rev_op_on_records(self):
        records = _parse_rows([_header_row("Gross Sales ($K)"), _data_row()], "CDE", 2024)
        assert all(r["rev_op_type"] == "Gross" for r in records)

    def test_second_header_changes_rev_op_for_subsequent_rows(self):
        rows = [
            _header_row("Gross Sales ($K)"),
            _data_row("APAC"),
            _header_row("Net Sales ($K)"),
            _data_row("EMEA"),
        ]
        records = _parse_rows(rows, "CDE", 2024)
        apac = [r for r in records if r["region"] == "APAC"]
        emea = [r for r in records if r["region"] == "EMEA"]
        assert apac and all(r["rev_op_type"] == "Gross" for r in apac)
        assert emea and all(r["rev_op_type"] == "Net" for r in emea)

    def test_product_and_year_embedded(self):
        records = _parse_rows([_header_row(), _data_row()], "PJ", 2025)
        assert all(r["product"] == "PJ" for r in records)
        assert all(r["year"] == 2025 for r in records)

    def test_section_type_label_row_does_not_produce_extra_records(self):
        rows_without = [_header_row(), _data_row()]
        rows_with = [_header_row(), (None, "Actual") + (None,) * 42, _data_row()]
        assert len(_parse_rows(rows_with, "CDE", 2024)) == len(_parse_rows(rows_without, "CDE", 2024))

    def test_multiple_data_rows_each_produce_records(self):
        rows = [_header_row(), _data_row("APAC"), _data_row("EMEA")]
        records = _parse_rows(rows, "CDE", 2024)
        regions = {r["region"] for r in records}
        assert "APAC" in regions
        assert "EMEA" in regions

    def test_only_header_no_data_returns_empty(self):
        assert _parse_rows([_header_row()], "CDE", 2024) == []


# ── load_excel (integration) ──────────────────────────────────────────────────

class TestLoadExcel:
    def test_invalid_filename_raises_value_error(self, tmp_path):
        from openpyxl import Workbook
        bad = tmp_path / "invalid_name.xlsx"
        Workbook().save(bad)
        with pytest.raises(ValueError, match="Cannot infer"):
            load_excel(bad)

    def test_missing_file_raises_file_not_found(self, tmp_path):
        with pytest.raises(FileNotFoundError):
            load_excel(tmp_path / "CDE biz status_2024_v1.xlsx")

    def test_empty_workbook_returns_empty_dataframe(self, tmp_path):
        from openpyxl import Workbook
        path = tmp_path / "CDE biz status_2024_v1.xlsx"
        Workbook().save(path)
        df = load_excel(path)
        assert df.empty

    def test_empty_workbook_has_output_columns(self, tmp_path):
        from openpyxl import Workbook
        path = tmp_path / "CDE biz status_2024_v1.xlsx"
        Workbook().save(path)
        df = load_excel(path)
        assert list(df.columns) == config.OUTPUT_COLUMNS

    def test_minimal_workbook_returns_records(self, tmp_path):
        from openpyxl import Workbook
        path = tmp_path / "CDE biz status_2024_v1.xlsx"
        wb = Workbook()
        ws = wb.active
        # Row 1: rev_op_type header in col B (1-indexed col 2)
        ws.cell(1, 2, "Gross Sales ($K)")
        # Row 2: data — region in col B, 12 Actual month values in cols C-N
        ws.cell(2, 2, "APAC")
        for i in range(12):
            ws.cell(2, 3 + i, float(i + 1))
        wb.save(path)
        df = load_excel(path)
        assert not df.empty
        assert set(config.OUTPUT_COLUMNS).issubset(set(df.columns))

    def test_minimal_workbook_product_and_year(self, tmp_path):
        from openpyxl import Workbook
        path = tmp_path / "PJ biz status_2025_v1.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.cell(1, 2, "Net Sales ($K)")
        ws.cell(2, 2, "APAC")
        for i in range(12):
            ws.cell(2, 3 + i, 100.0)
        wb.save(path)
        df = load_excel(path)
        assert (df["product"] == "PJ").all()
        assert (df["year"] == 2025).all()

    def test_real_xlsx_files_parse_correctly(self):
        xlsx_files = sorted(config.RAW_DATA_DIR.glob("*.xlsx"))
        if not xlsx_files:
            pytest.skip("No .xlsx files in raw_data/")
        for path in xlsx_files:
            df = load_excel(path)
            assert not df.empty, f"Empty result for {path.name}"
            assert set(config.OUTPUT_COLUMNS).issubset(set(df.columns)), path.name
            assert df["product"].notna().all()
            assert df["amount"].dtype.kind == "f"  # raw float before transform
